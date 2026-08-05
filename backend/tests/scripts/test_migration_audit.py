import sys
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parents[2] / "scripts"
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from migration_audit import MigrationAudit, write_audit_workbook  # noqa: E402


def test_audit_deduplicates_by_source_issue_and_field():
    audit = MigrationAudit()
    for target_value in ("first", "second"):
        audit.record(
            source_table="recording", source_id=1, target_table="media", target_id=1,
            issue_type="field_mismatch", severity="error", field_name="filename",
            source_value="source.wav", target_value=target_value, reason="Different values.",
            recommended_action="Repair the target value.",
        )

    assert len(audit.issues) == 1
    assert audit.issues[0].target_value == "first"


def test_write_audit_workbook_uses_source_table_sheets_only(tmp_path):
    audit = MigrationAudit()
    audit.record(
        source_table="recording", source_id=5, target_table="media", target_id=5,
        issue_type="file_missing", severity="error", field_name="filename",
        source_value="missing.wav", reason="Audio file missing.",
        recommended_action="Restore the file.",
    )
    audit.record(
        source_table="queue", source_id=7, target_table="queue", target_id=7,
        issue_type="unsupported_value", severity="warning", field_name="status",
        source_value=99, target_value=3, reason="Unknown status.",
        recommended_action="Review the target status.",
    )
    output = tmp_path / "audit.xlsx"

    assert write_audit_workbook(audit, output) is True

    workbook = load_workbook(output)
    assert workbook.sheetnames == ["queue", "recording"]
    worksheet = workbook["recording"]
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:K2"
    assert worksheet["A2"].value == "recording"
    assert worksheet["E2"].value == "file_missing"


def test_write_audit_workbook_skips_empty_audit(tmp_path):
    output = tmp_path / "audit.xlsx"

    assert write_audit_workbook(MigrationAudit(), output) is False
    assert not output.exists()
