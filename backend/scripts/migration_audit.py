"""Disk-backed migration audit records and streaming XLSX export helpers."""

from __future__ import annotations

import json
import os
import tempfile
from collections.abc import Iterator
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

AUDIT_COLUMNS = [
    "source_table", "source_id", "target_table", "target_id", "issue_type", "severity",
    "field_name", "source_value", "target_value", "reason", "recommended_action",
]
EXCEL_MAX_DATA_ROWS = 1_048_575


@dataclass(frozen=True)
class MigrationAuditIssue:
    source_table: str
    source_id: str
    target_table: str | None
    target_id: str | None
    issue_type: str
    severity: str
    field_name: str | None
    source_value: Any
    target_value: Any
    reason: str
    recommended_action: str


class MigrationAudit:
    """Append row-level audit issues to disk without retaining their values in memory."""

    def __init__(self) -> None:
        fd, name = tempfile.mkstemp(prefix="ecosignal-migration-audit-", suffix=".jsonl")
        os.close(fd)
        Path(name).touch(exist_ok=True)
        self._path = Path(name)
        self._keys: set[tuple[str, str, str, str | None]] = set()
        self.count = 0

    @property
    def issues(self) -> list[MigrationAuditIssue]:
        return list(self.iter_issues())

    def iter_issues(self) -> Iterator[MigrationAuditIssue]:
        with self._path.open(encoding="utf-8") as handle:
            for line in handle:
                yield MigrationAuditIssue(**json.loads(line))

    def add(self, issue: MigrationAuditIssue) -> None:
        key = (issue.source_table, issue.source_id, issue.issue_type, issue.field_name)
        if key in self._keys:
            return
        self._keys.add(key)
        with self._path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(asdict(issue), default=str, ensure_ascii=False) + "\n")
        self.count += 1

    def record(self, **kwargs: Any) -> None:
        self.add(MigrationAuditIssue(
            source_table=kwargs["source_table"], source_id=str(kwargs["source_id"]),
            target_table=kwargs.get("target_table"),
            target_id=None if kwargs.get("target_id") is None else str(kwargs["target_id"]),
            issue_type=kwargs["issue_type"], severity=kwargs["severity"],
            field_name=kwargs.get("field_name"), source_value=kwargs.get("source_value"),
            target_value=kwargs.get("target_value"), reason=kwargs["reason"],
            recommended_action=kwargs["recommended_action"],
        ))


def write_audit_workbook(audit: MigrationAudit, output_path: Path) -> bool:
    """Write source-table worksheets from the disk-backed audit stream."""
    if audit.count == 0:
        return False

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    source_tables = sorted({issue.source_table for issue in audit.iter_issues()})
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for source_table in source_tables:
        sheet_index = 1
        data_rows = 0
        worksheet = _new_sheet(workbook, source_table, sheet_index, header_fill, header_font)
        for issue in audit.iter_issues():
            if issue.source_table != source_table:
                continue
            if data_rows == EXCEL_MAX_DATA_ROWS:
                worksheet.auto_filter.ref = f"A1:K{data_rows + 1}"
                sheet_index += 1
                data_rows = 0
                worksheet = _new_sheet(workbook, source_table, sheet_index, header_fill, header_font)
            values = asdict(issue)
            worksheet.append([_excel_value(values[column]) for column in AUDIT_COLUMNS])
            data_rows += 1
        worksheet.auto_filter.ref = f"A1:K{data_rows + 1}"
    workbook.save(output_path)
    return True


def _new_sheet(workbook, source_table: str, index: int, header_fill, header_font):
    from openpyxl.cell import WriteOnlyCell

    title = source_table if index == 1 else f"{source_table}_{index}"
    worksheet = workbook.create_sheet(title=title[:31])
    worksheet.freeze_panes = "A2"
    header = []
    for name in AUDIT_COLUMNS:
        cell = WriteOnlyCell(worksheet, value=name)
        cell.fill = header_fill
        cell.font = header_font
        header.append(cell)
    worksheet.append(header)
    return worksheet


def _excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
